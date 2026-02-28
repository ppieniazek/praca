from datetime import datetime

from datastar_py import ServerSentEventGenerator as SSE
from datastar_py.django import DatastarResponse
from django.contrib import messages
from django.db import transaction
from django.db.models import DecimalField, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect

from business.models import BonusDay, Payroll, WalletTransaction, Worker, WorkLog
from business.views.utils import (
    get_toast_event,
    get_user_org,
    is_owner,
    render_template,
)

import os
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import Http404

class PayrollPDF(FPDF):
    """Prosty generator PDF dla listy płac."""
    def __init__(self, org_name, period_str):
        super().__init__()
        self.org_name = org_name
        self.period_str = period_str
        
        font_dir = "/usr/share/fonts/liberation-sans-fonts/"
        if not os.path.exists(font_dir):
            font_dir = "/usr/share/fonts/liberation/"
            
        self.font_name = "helvetica"
        if os.path.exists(font_dir):
            fonts = {
                "": "LiberationSans-Regular.ttf",
                "B": "LiberationSans-Bold.ttf",
                "I": "LiberationSans-Italic.ttf",
                "BI": "LiberationSans-BoldItalic.ttf",
            }
            loaded = False
            for style, filename in fonts.items():
                path = os.path.join(font_dir, filename)
                if os.path.exists(path):
                    self.add_font("Liberation", style, path)
                    loaded = True
            if loaded:
                self.font_name = "Liberation"

    def header(self):
        self.set_font(self.font_name, "B", 16)
        self.cell(0, 10, f"Lista płac - {self.period_str}", align="C")
        self.ln(10)
        self.set_font(self.font_name, "", 12)
        self.cell(0, 10, self.org_name, align="C")
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font(self.font_name, "I", 8)
        self.cell(0, 10, f"Strona {self.page_no()}", align="C")

def payroll_export_pdf_view(request: HttpRequest):
    """Eksportuje zamknięte wypłaty do formatu PDF."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return HttpResponse(status=403)
    
    organization = get_user_org(request.user)
    year, month = _get_year_month(request)
    
    payrolls = Payroll.objects.filter(
        organization=organization,
        year=year,
        month=month,
        status=Payroll.Status.CLOSED
    ).select_related("worker")
    
    if not payrolls.exists():
        raise Http404("Brak zamkniętych wypłat dla wybranego miesiąca.")
        
    pdf = PayrollPDF(organization.name, f"{month:02d}/{year}")
    pdf.add_page()
    
    pdf.set_font(pdf.font_name, "B", 10)
    cols = [
        ("Pracownik", 55), 
        ("Stawka", 25), 
        ("Godziny", 25), 
        ("Bonusy", 25), 
        ("Zaliczki", 25), 
        ("Razem", 25)
    ]
    for col_name, width in cols:
        pdf.cell(width, 10, col_name, border=1, align="C")
    pdf.ln()
    
    pdf.set_font(pdf.font_name, "", 10)
    total_h = 0
    total_bonus = 0
    total_adv = 0
    total_net = 0
    for p in payrolls:
        name = f"{p.worker.first_name} {p.worker.last_name}"
        pdf.cell(55, 10, name, border=1)
        pdf.cell(25, 10, f"{p.hourly_rate_snapshot}", border=1, align="R")
        pdf.cell(25, 10, f"{p.total_hours}", border=1, align="R")
        pdf.cell(25, 10, f"{p.bonuses:.2f}", border=1, align="R")
        pdf.cell(25, 10, f"{p.advances_deducted:.2f}", border=1, align="R")
        pdf.cell(25, 10, f"{p.net_pay:.2f}", border=1, align="R")
        pdf.ln()
        total_h += p.total_hours
        total_bonus += p.bonuses
        total_adv += p.advances_deducted
        total_net += p.net_pay
        
    pdf.set_font(pdf.font_name, "B", 10)
    pdf.cell(80, 10, "SUMA", border=1, align="R")
    pdf.cell(25, 10, f"{total_h}", border=1, align="R")
    pdf.cell(25, 10, f"{total_bonus:.2f}", border=1, align="R")
    pdf.cell(25, 10, f"{total_adv:.2f}", border=1, align="R")
    pdf.cell(25, 10, f"{total_net:.2f}", border=1, align="R")
    
    response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="wyplaty_{month:02d}_{year}.pdf"'
    return response

def payroll_export_excel_view(request: HttpRequest):
    """Eksportuje zamknięte wypłaty do formatu EXCEL (XLSX)."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return HttpResponse(status=403)
        
    organization = get_user_org(request.user)
    year, month = _get_year_month(request)
    
    payrolls = Payroll.objects.filter(
        organization=organization,
        year=year,
        month=month,
        status=Payroll.Status.CLOSED
    ).select_related("worker")
    
    if not payrolls.exists():
        raise Http404("Brak zamkniętych wypłat dla wybranego miesiąca.")
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista płac"
    
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Lista płac - {month:02d}/{year} ({organization.name})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    headers = ["Pracownik", "Stawka", "Godziny", "Bonusy", "Zaliczki", "Razem"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    total_h = 0
    total_bonus = 0
    total_adv = 0
    total_net = 0
    for p in payrolls:
        ws.append([
            f"{p.worker.first_name} {p.worker.last_name}",
            p.hourly_rate_snapshot,
            p.total_hours,
            float(p.bonuses),
            float(p.advances_deducted),
            float(p.net_pay)
        ])
        total_h += p.total_hours
        total_bonus += p.bonuses
        total_adv += p.advances_deducted
        total_net += p.net_pay
        
    last_row = ws.max_row + 1
    ws.append(["SUMA", "", total_h, float(total_bonus), float(total_adv), float(total_net)])
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
    
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="wyplaty_{month:02d}_{year}.xlsx"'
    wb.save(response)
    return response

@transaction.atomic
def bonus_day_manage_view(request: HttpRequest):
    """Zarządzanie bonusami (dniówkami premium) dla danej organizacji."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return DatastarResponse(SSE.redirect("/login/"))

    organization = get_user_org(request.user)
    year, month = _get_year_month(request)

    is_closed = Payroll.objects.filter(
        organization=organization, year=year, month=month, status=Payroll.Status.CLOSED
    ).exists()

    if request.method == "POST":
        if is_closed:
            messages.error(request, "Nie można edytować bonusów w zamkniętym miesiącu.")
            return DatastarResponse(get_toast_event(request))

        action = request.POST.get("action")
        if action == "add":
            try:
                date_str = request.POST.get("date")
                amount = int(request.POST.get("amount", 0))
                description = request.POST.get("description", "")
                bonus_date = datetime.strptime(date_str, "%Y-%m-%d").date()

                if bonus_date.year != year or bonus_date.month != month:
                    messages.error(request, "Data musi mieścić się w wybranym miesiącu.")
                else:
                    BonusDay.objects.update_or_create(
                        organization=organization,
                        date=bonus_date,
                        defaults={"amount": amount, "description": description},
                    )
                    messages.success(request, f"Dodano bonus dla dnia {bonus_date}.")
            except (ValueError, TypeError):
                messages.error(request, "Nieprawidłowe dane formularza.")
        elif action == "delete":
            bonus_id = request.POST.get("bonus_id")
            BonusDay.objects.filter(id=bonus_id, organization=organization).delete()
            messages.warning(request, "Usunięto bonus.")

    bonus_days = BonusDay.objects.filter(
        organization=organization, date__year=year, date__month=month
    ).order_by("date")

    context = {
        "bonus_days": bonus_days,
        "current_year": year,
        "current_month": month,
        "is_closed": is_closed,
    }

    rendered = render_template(
        "business/timesheet_grid.html#bonus_day_manage_modal", context, request
    )
    return DatastarResponse(
        [
            SSE.patch_elements(rendered, selector="#modal-content"),
            SSE.patch_signals({"is_modal_open": True}),
            get_toast_event(request) if request.method == "POST" else None,
        ]
    )

def _get_year_month(request) -> tuple[int, int]:
    now = datetime.now()
    try:
        year = int(request.GET.get("year", now.year))
        month = int(request.GET.get("month", now.month))

        if month < 1:
            month, year = 12, year - 1
        elif month > 12:
            month, year = 1, year + 1
    except (ValueError, TypeError):
        year, month = now.year, now.month
    return year, month


def get_payrolls_for_month(organization, year, month):
    return list(
        Payroll.objects.filter(
            organization=organization, year=year, month=month
        ).select_related("worker")
    )


def get_payroll_stats(payrolls):
    total_earned = sum(p.gross_pay for p in payrolls)
    total_bonuses = sum(p.bonuses for p in payrolls)
    total_payout = sum(p.net_pay for p in payrolls)
    total_advances = sum(p.advances_deducted for p in payrolls)
    has_drafts = any(p.status == Payroll.Status.DRAFT for p in payrolls)
    has_closed = any(p.status == Payroll.Status.CLOSED for p in payrolls)

    status = "EMPTY"
    if has_closed and not has_drafts:
        status = "CLOSED"
    elif has_drafts:
        status = "DRAFT"

    return {
        "total_earned": total_earned,
        "total_bonuses": total_bonuses,
        "total_payout": total_payout,
        "total_advances": total_advances,
        "status": status,
    }


def payroll_list_view(request: HttpRequest):
    """Zarządzanie wypłatami i podgląd na dany miesiąc."""
    if not request.user.is_authenticated or not is_owner(request.user):
        messages.error(request, "Brak dostępu do widoku wypłat.")
        return redirect("core:dashboard")

    organization = get_user_org(request.user)
    year, month = _get_year_month(request)

    payrolls = get_payrolls_for_month(organization, year, month)
    stats = get_payroll_stats(payrolls)

    context = {
        "payrolls": payrolls,
        "current_year": year,
        "current_month": month,
        "stats": stats,
    }

    if "Datastar-Request" in request.headers:
        rendered = render_template(
            "business/payroll_list.html#payroll_content", context, request
        )
        return DatastarResponse(
            [
                SSE.patch_elements(
                    rendered, selector="#payroll-container", mode="inner"
                ),
                get_toast_event(request),
            ]
        )

    return HttpResponse(render_template("business/payroll_list.html", context, request))


@transaction.atomic
def payroll_generate_month_view(request: HttpRequest):
    """Generuje lub przelicza wypłaty (DRAFT) dla wszystkich pracowników za dany miesiąc."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return DatastarResponse(SSE.redirect("/login/"))

    if request.method != "POST":
        return HttpResponse(status=405)

    organization = get_user_org(request.user)
    try:
        year = int(request.GET.get("year"))
        month = int(request.GET.get("month"))
    except (ValueError, TypeError):
        messages.error(request, "Nieprawidłowy miesiąc lub rok.")
        return DatastarResponse(get_toast_event(request))

    workers_with_stats = (
        Worker.objects.filter(organization=organization)
        .annotate(
            total_hours=Coalesce(
                Sum(
                    "work_logs__hours",
                    filter=Q(work_logs__date__year=year, work_logs__date__month=month),
                ),
                0,
                output_field=DecimalField(),
            ),
            total_advances=Coalesce(
                Sum(
                    "advances__amount",
                    filter=Q(
                        advances__organization=organization,
                        advances__type=WalletTransaction.Type.ADVANCE,
                        advances__date__year=year,
                        advances__date__month=month,
                    ),
                ),
                0,
                output_field=DecimalField(),
            ),
        )
        .filter(Q(total_hours__gt=0) | Q(total_advances__gt=0))
        .distinct()
    )

    generated_count = 0
    skipped_count = 0

    bonus_days = BonusDay.objects.filter(
        organization=organization, date__year=year, date__month=month
    )
    bonus_map = {bd.date: bd.amount for bd in bonus_days}
    bonus_dates = list(bonus_map.keys())

    existing_payrolls = {
        p.worker_id: p
        for p in Payroll.objects.filter(organization=organization, year=year, month=month)
    }

    worker_bonuses = {}
    if bonus_dates:
        worker_ids = [w.id for w in workers_with_stats]
        bonus_logs = WorkLog.objects.filter(
            organization=organization,
            worker_id__in=worker_ids,
            date__in=bonus_dates,
            hours__gt=0
        ).values("worker_id", "date")
        
        for log in bonus_logs:
            wid = log["worker_id"]
            d = log["date"]
            worker_bonuses[wid] = worker_bonuses.get(wid, 0) + bonus_map.get(d, 0)

    for worker in workers_with_stats:
        existing = existing_payrolls.get(worker.id)

        if existing and existing.status == Payroll.Status.CLOSED:
            skipped_count += 1
            continue

        worker_bonus_sum = worker_bonuses.get(worker.id, 0)

        hourly_rate = worker.hourly_rate
        gross_pay = (worker.total_hours * hourly_rate) + worker_bonus_sum
        net_pay = gross_pay - worker.total_advances

        if existing:
            existing.total_hours = worker.total_hours
            existing.hourly_rate_snapshot = hourly_rate
            existing.bonuses = worker_bonus_sum
            existing.gross_pay = gross_pay
            existing.advances_deducted = worker.total_advances
            existing.net_pay = net_pay
            existing.save()
        else:
            Payroll.objects.create(
                organization=organization,
                worker=worker,
                year=year,
                month=month,
                status=Payroll.Status.DRAFT,
                total_hours=worker.total_hours,
                hourly_rate_snapshot=hourly_rate,
                bonuses=worker_bonus_sum,
                gross_pay=gross_pay,
                advances_deducted=worker.total_advances,
                net_pay=net_pay,
            )
        generated_count += 1

    messages.success(
        request,
        f"Przeliczono listę płac za {month:02d}/{year} ({generated_count} wpisów).",
    )
    if skipped_count > 0:
        messages.warning(
            request,
            f"Pominięto {skipped_count} wpisów, ponieważ były już zamknięte (CLOSED).",
        )

    payrolls = get_payrolls_for_month(organization, year, month)
    stats = get_payroll_stats(payrolls)
    context = {
        "payrolls": payrolls,
        "current_year": year,
        "current_month": month,
        "stats": stats,
    }

    rendered = render_template(
        "business/payroll_list.html#payroll_content", context, request
    )
    return DatastarResponse(
        [
            SSE.patch_elements(rendered, selector="#payroll-container", mode="inner"),
            get_toast_event(request),
        ]
    )


@transaction.atomic
def payroll_close_month_view(request: HttpRequest):
    """Zamyka wszystkie wypłaty o statusie DRAFT dla danego miesiąca."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return DatastarResponse(SSE.redirect("/login/"))

    if request.method != "POST":
        return HttpResponse(status=405)

    organization = get_user_org(request.user)
    try:
        year = int(request.GET.get("year"))
        month = int(request.GET.get("month"))
    except (ValueError, TypeError):
        messages.error(request, "Nieprawidłowy miesiąc lub rok.")
        return DatastarResponse(get_toast_event(request))

    drafts = Payroll.objects.filter(
        organization=organization, year=year, month=month, status=Payroll.Status.DRAFT
    )
    count = drafts.update(status=Payroll.Status.CLOSED)

    messages.success(
        request,
        f"Zamknięto miesiąc {month:02d}/{year}. Edycja godzin w tym miesiącu została zablokowana dla {count} pracowników.",
    )

    payrolls = get_payrolls_for_month(organization, year, month)
    stats = get_payroll_stats(payrolls)
    context = {
        "payrolls": payrolls,
        "current_year": year,
        "current_month": month,
        "stats": stats,
    }

    rendered = render_template(
        "business/payroll_list.html#payroll_content", context, request
    )
    return DatastarResponse(
        [
            SSE.patch_elements(rendered, selector="#payroll-container", mode="inner"),
            get_toast_event(request),
        ]
    )


@transaction.atomic
def payroll_reopen_month_view(request: HttpRequest):
    """Przewraca wypłaty zamkniętego miesiąca z powrotem do statusu DRAFT."""
    if not request.user.is_authenticated or not is_owner(request.user):
        return DatastarResponse(SSE.redirect("/login/"))

    if request.method != "POST":
        return HttpResponse(status=405)

    organization = get_user_org(request.user)
    try:
        year = int(request.GET.get("year"))
        month = int(request.GET.get("month"))
    except (ValueError, TypeError):
        messages.error(request, "Nieprawidłowy miesiąc lub rok.")
        return DatastarResponse(get_toast_event(request))

    closed = Payroll.objects.filter(
        organization=organization, year=year, month=month, status=Payroll.Status.CLOSED
    )
    count = closed.update(status=Payroll.Status.DRAFT)

    messages.warning(
        request,
        f"Miesiąc {month:02d}/{year} został otwarty ({count} wypłat do poprawy). Blokada edycji Timesheet zdjęta.",
    )

    payrolls = get_payrolls_for_month(organization, year, month)
    stats = get_payroll_stats(payrolls)
    context = {
        "payrolls": payrolls,
        "current_year": year,
        "current_month": month,
        "stats": stats,
    }

    rendered = render_template(
        "business/payroll_list.html#payroll_content", context, request
    )
    return DatastarResponse(
        [
            SSE.patch_elements(rendered, selector="#payroll-container", mode="inner"),
            get_toast_event(request),
        ]
    )
